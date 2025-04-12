import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference, PieChart, LineChart

def generate_excel_dashboard(df, output_path="violations_dashboard.xlsx"):
    df = df.copy()
    df["full_address"] = df["house_number"].astype(str) + " " + df["street"]

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "violations_all"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws_data[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    for col in ws_data.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws_data.column_dimensions[col[0].column_letter].width = max_len + 2

    ws_data.freeze_panes = "A2"

    # Dashboard Sheet
    ws_dash = wb.create_sheet("dashboard")
    ws_dash["A1"] = "DOB Violation Insights Dashboard"
    ws_dash["A1"].font = Font(size=14, bold=True)

    df["month"] = pd.to_datetime(df["issue_date"], errors='coerce').dt.to_period('M').astype(str)

    # Top streets
    top_streets = df["street"].value_counts().head(10).reset_index()
    top_streets.columns = ["Street", "Count"]
    start_row = 3
    for row in dataframe_to_rows(top_streets, index=False, header=True):
        ws_dash.append(row)
    bar1 = BarChart()
    bar1.title = "Top 10 Streets by Violations"
    data = Reference(ws_dash, min_col=2, min_row=start_row+1, max_row=start_row+10)
    labels = Reference(ws_dash, min_col=1, min_row=start_row+2, max_row=start_row+11)
    bar1.add_data(data, titles_from_data=True)
    bar1.set_categories(labels)
    ws_dash.add_chart(bar1, "E3")

    # Top addresses
    top_addrs = df["full_address"].value_counts().head(10).reset_index()
    top_addrs.columns = ["Address", "Count"]
    start_row = start_row + len(top_streets) + 5
    for row in dataframe_to_rows(top_addrs, index=False, header=True):
        ws_dash.append(row)
    bar2 = BarChart()
    bar2.title = "Top 10 Repeat Violation Addresses"
    data = Reference(ws_dash, min_col=2, min_row=start_row+1, max_row=start_row+10)
    labels = Reference(ws_dash, min_col=1, min_row=start_row+2, max_row=start_row+11)
    bar2.add_data(data, titles_from_data=True)
    bar2.set_categories(labels)
    ws_dash.add_chart(bar2, "E18")

    # Violation Types
    types = df["violation_type"].value_counts().reset_index()
    types.columns = ["Type", "Count"]
    start_row = start_row + len(top_addrs) + 5
    for row in dataframe_to_rows(types, index=False, header=True):
        ws_dash.append(row)
    bar3 = BarChart()
    bar3.title = "Top Violation Types"
    data = Reference(ws_dash, min_col=2, min_row=start_row+1, max_row=start_row+len(types))
    labels = Reference(ws_dash, min_col=1, min_row=start_row+2, max_row=start_row+len(types))
    bar3.add_data(data, titles_from_data=True)
    bar3.set_categories(labels)
    ws_dash.add_chart(bar3, "E34")

    # Over Time
    monthly = df["month"].value_counts().sort_index().reset_index()
    monthly.columns = ["Month", "Count"]
    start_row = start_row + len(types) + 5
    for row in dataframe_to_rows(monthly, index=False, header=True):
        ws_dash.append(row)
    line = LineChart()
    line.title = "Violations Over Time"
    data = Reference(ws_dash, min_col=2, min_row=start_row+1, max_row=start_row+len(monthly))
    labels = Reference(ws_dash, min_col=1, min_row=start_row+2, max_row=start_row+len(monthly))
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    ws_dash.add_chart(line, "E45")

    wb.save(output_path)

def generate_excel(df, output_path='violations.xlsx'):
    return generate_excel_dashboard(df, output_path)
